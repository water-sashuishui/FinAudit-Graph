from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path("data/demo_inputs/realistic_business_financials.xlsx")


def main() -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = {
        "Cover": [
            ["财务审计资料包"],
            ["被审计单位", "远航智能制造有限公司"],
            ["统一社会信用代码", "91310000MA1K202406"],
            ["报告年度", "2024"],
            ["资料范围", "年度财务报表、收入明细、应收账款账龄、现金流量资料"],
            ["编制说明", "本资料为模拟业务数据，用于 FinAudit-Graph 演示。"],
        ],
        "IncomeStatement": [
            ["编制单位：远航智能制造有限公司"],
            ["单位：万元"],
            ["项目", "本期金额", "上期金额", "变动说明"],
            ["一、营业收入", 12800, 10000, "第四季度大客户集中验收"],
            ["减：营业成本", 7680, 6800, "核心材料采购成本上升"],
            ["税金及附加", 96, 82, ""],
            ["销售费用", 960, 720, "新增区域销售团队"],
            ["管理费用", 1180, 1020, "研发管理投入增加"],
            ["研发费用", 1450, 1200, "智能产线项目投入"],
            ["财务费用", 260, 180, "短期借款增加"],
            ["二、营业利润", 1174, 998, ""],
            ["三、利润总额", 1160, 990, ""],
            ["四、净利润", 986, 842, ""],
        ],
        "CashFlow": [
            ["编制单位：远航智能制造有限公司"],
            ["单位：万元"],
            ["项目", "本期金额", "上期金额", "备注"],
            ["销售商品、提供劳务收到的现金", 9300, 8800, "回款增长低于收入增长"],
            ["购买商品、接受劳务支付的现金", 6900, 6100, "供应商账期缩短"],
            ["支付给职工以及为职工支付的现金", 1180, 960, ""],
            ["支付的各项税费", 500, 420, ""],
            ["经营活动产生的现金流量净额", 720, 1000, "收入增长但经营现金流下降"],
            ["投资活动产生的现金流量净额", -1850, -900, "新增生产线设备"],
            ["筹资活动产生的现金流量净额", 1600, 300, "新增短期借款"],
        ],
        "BalanceSheet": [
            ["编制单位：远航智能制造有限公司"],
            ["单位：万元"],
            ["项目", "期末余额", "期初余额", "备注"],
            ["货币资金", 2380, 1910, ""],
            ["应收票据", 620, 480, ""],
            ["应收账款", 4620, 3300, "大客户赊销余额上升"],
            ["预付款项", 880, 510, "预付材料款增加"],
            ["存货", 3900, 3100, "备货增加"],
            ["固定资产", 8200, 6700, "新增产线"],
            ["短期借款", 3200, 1500, "流动资金贷款"],
            ["应付账款", 2800, 2300, ""],
            ["所有者权益合计", 10200, 9300, ""],
        ],
        "RevenueDetail": [
            ["客户名称", "是否关联方", "合同编号", "收入金额", "验收日期", "回款金额", "备注"],
            ["华东轨交装备集团", "否", "HT-2024-001", 2600, "2024-03-28", 2100, "分批验收"],
            ["远航供应链科技有限公司", "是", "HT-2024-087", 1850, "2024-12-29", 220, "期末集中确认"],
            ["北方智能制造有限公司", "否", "HT-2024-051", 1430, "2024-09-18", 980, ""],
            ["新港自动化工程有限公司", "否", "HT-2024-099", 1260, "2024-12-30", 100, "期后回款待跟踪"],
            ["西南能源设备有限公司", "否", "HT-2024-034", 1180, "2024-06-20", 960, ""],
        ],
        "ReceivableAging": [
            ["客户名称", "期末余额", "1年以内", "1-2年", "2年以上", "坏账准备", "备注"],
            ["远航供应链科技有限公司", 1630, 1450, 180, 0, 45, "疑似关联方余额较高"],
            ["华东轨交装备集团", 980, 980, 0, 0, 20, ""],
            ["新港自动化工程有限公司", 760, 760, 0, 0, 18, "期末新增"],
            ["北方智能制造有限公司", 520, 480, 40, 0, 16, ""],
            ["其他客户", 730, 690, 40, 0, 22, ""],
        ],
    }

    for title, rows in sheets.items():
        worksheet = workbook.create_sheet(title)
        for row in rows:
            worksheet.append(row)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
        for col in range(1, worksheet.max_column + 1):
            width = max(
                len(str(worksheet.cell(row=row, column=col).value or ""))
                for row in range(1, worksheet.max_row + 1)
            )
            worksheet.column_dimensions[get_column_letter(col)].width = min(max(width + 4, 12), 34)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH.resolve())


if __name__ == "__main__":
    main()
